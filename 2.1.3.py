import csv
import openpyxl
import numpy as np
import matplotlib.pyplot as mat
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter

from jinja2 import Environment, FileSystemLoader
import pathlib
import pdfkit


class Vacancy:
    currency_conv_dic = {
        "AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
        "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055,
    }

    def __init__(self, vacancy):
        self.name = vacancy['name']
        self.salary_currency = vacancy['salary_currency']
        self.salary_from = int(float(vacancy['salary_from']))
        self.salary_to = int(float(vacancy['salary_to']))
        self.area_name = vacancy['area_name']
        self.year = int(vacancy['published_at'][:4])
        self.salary_average = self.currency_conv_dic[self.salary_currency] * (self.salary_from + self.salary_to) / 2


class DataSet:
    def __init__(self, name, vacancy):
        self.file_name = name
        self.vacancy_name = vacancy

    @staticmethod
    def get_average_sum_by_dic(dic):
        temp = {}
        for k, v in dic.items():
            temp[k] = int(sum(v) / len(v))
        return temp

    @staticmethod
    def increase_value(dic, k, v):
        if k in dic:
            dic[k] += v
        else:
            dic[k] = v

    def csv_reader(self):
        with open(self.file_name, encoding='utf-8-sig') as file:
            csv_file = csv.reader(file)
            fields = next(csv_file)
            amount_of_fields = len(fields)
            for element in csv_file:
                if '' not in element and len(element) == amount_of_fields:
                    yield dict(zip(fields, element))

    def get_statistic(self):
        salary = {}
        area_salary = {}
        salary_of_vac = {}

        counter, name_num_vac, num_vac, salary_of_vac = self.statistics_helper(area_salary, 0, salary, salary_of_vac)

        stats = self.get_average_sum_by_dic(salary)
        s1 = self.get_average_sum_by_dic(salary_of_vac)
        s3 = self.get_average_sum_by_dic(area_salary)

        s4 = {}
        for year, sal in area_salary.items():
            s4[year] = round(len(sal) / counter, 4)
        s4 = list(filter(lambda a: a[-1] >= 0.01, [(k, v) for k, v in s4.items()]))
        s4.sort(key=lambda a: a[-1], reverse=True)
        s5 = s4.copy()
        s4 = dict(s4)
        s3 = list(filter(lambda a: a[0] in list(s4.keys()), [(k, v) for k, v in s3.items()]))
        s3.sort(key=lambda a: a[-1], reverse=True)
        s3 = dict(s3[:10])
        s5 = dict(s5[:10])

        return stats, num_vac, s1, name_num_vac, s3, s5

    def statistics_helper(self, area_salary, counter, salary, salary_of_vac):
        for vac_dic in self.csv_reader():
            vac = Vacancy(vac_dic)
            self.increase_value(salary, vac.year, [vac.salary_average])
            if vac.name.find(self.vacancy_name) != -1:
                self.increase_value(salary_of_vac, vac.year, [vac.salary_average])
            self.increase_value(area_salary, vac.area_name, [vac.salary_average])
            counter += 1
        num_vac = dict([(k, len(v)) for k, v in salary.items()])
        name_num_vac = dict([(k, len(v)) for k, v in salary_of_vac.items()])
        if not salary_of_vac:
            salary_of_vac = dict([(k, [0]) for k, v in salary.items()])
            name_num_vac = dict([(k, 0) for k, v in num_vac.items()])
        return counter, name_num_vac, num_vac, salary_of_vac

    @staticmethod
    def print_statistic(s1, s2, s3, s4, s5, s6):
        print('Динамика уровня зарплат по годам: {0}'.format(s1))
        print('Динамика количества вакансий по годам: {0}'.format(s2))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(s3))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(s4))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(s5))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(s6))


class InputConnect:
    def __init__(self):
        self.name = input('Введите название файла: ')
        self.vacancy = input('Введите название профессии: ')

        ds = DataSet(self.name, self.vacancy)
        s1, s2, s3, s4, s5, s6 = ds.get_statistic()
        ds.print_statistic(s1, s2, s3, s4, s5, s6)
        rep = Report(self.vacancy, s1, s2, s3, s4, s5, s6)
        rep.generate_excel()
        rep.generate_image()
        rep.save('report.xlsx')
        rep.generate_pdf()


class Report:
    def __init__(self, vacancy, s1, s2, s3, s4, s5, s6):
        self.s1, self.s2, self.s3, self.s4, self.s5, self.s6 = s1, s2, s3, s4, s5, s6
        self.vacancy = vacancy
        self.book = openpyxl.Workbook()

    def generate_excel(self):
        first = self.book.active
        first.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancy, 'Количество вакансий',
                      'Количество вакансий - ' + self.vacancy])
        first.title = 'Статистика по годам'
        self.year_creator(first)
        data_field = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancy, ' Количество вакансий',
                       ' Количество вакансий - ' + self.vacancy]]
        widths = []
        letters = ["ABCDE", "ABDE"]
        for element in data_field:
            for i, j in enumerate(element):
                if len(widths) > i:
                    if len(j) > widths[i]:
                        widths[i] = len(j)
                else:
                    widths += [len(j)]

        for i, widths in enumerate(widths, 1):
            first.column_dimensions[get_column_letter(i)].width = widths + 2
        data_field = [['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']]
        for (first_city, first_value), (second_city, second_value) in zip(self.s5.items(), self.s6.items()):
            data_field.append([first_city, first_value, '', second_city, second_value])
        second = self.book.create_sheet('Статистика по городам')
        for element in data_field:
            second.append(element)

        widths = []
        for element in data_field:
            for i, j in enumerate(element):
                j = str(j)
                if len(widths) > i:
                    if len(j) > widths[i]:
                        widths[i] = len(j)
                else:
                    widths += [len(j)]

        for i, widths in enumerate(widths, 1):
            second.column_dimensions[get_column_letter(i)].width = widths + 2

        for e in letters[0]:
            first[e + '1'].font = Font(bold=True)
            second[e + '1'].font = Font(bold=True)

        for i, not_need in enumerate(self.s5):
            second['E' + str(i + 2)].number_format = '0.00%'

        temp_border = Border(left=Side(style='thin', color='00000000'), right=Side(style='thin', color='00000000'),
                             top=Side(style='thin', color='00000000'), bottom=Side(style='thin', color='00000000'))

        for element in range(len(data_field)):
            for e in letters[1]:
                second[e + str(element + 1)].border = temp_border

        for element, not_need in enumerate(self.s1):
            for e in letters[0]:
                first[e + str(element + 1)].border = temp_border

    def year_creator(self, first):
        for year in self.s1.keys():
            first.append([year, self.s1[year], self.s3[year], self.s2[year], self.s4[year]])

    def generate_image(self):
        not_need, ((a, b), (c, d)) = mat.subplots(ncols=2, nrows=2)
        self.first_graph(a)
        self.second_graph(b)
        self.third_graph(c)
        self.fourth_graph(d)

        mat.tight_layout()
        mat.savefig('graph.png')

    def fourth_graph(self, d):
        d.set_title('Доля вакансий по городам', fontdict={'fontsize': 8})
        temp = 1 - sum([value for value in self.s6.values()])
        d.pie(list(self.s6.values()) + [temp], labels=list(self.s6.keys()) + ['Другие'], textprops={'fontsize': 6})

    def third_graph(self, c):
        c.set_title('Уровень зарплат по городам', fontdict={'fontsize': 8})
        c.barh(list([str(a).replace(' ', '\n').replace('-', '-\n') for a in reversed(list(self.s5.keys()))]),
               list(reversed(list(self.s5.values()))), color='blue', height=0.5, align='center')
        c.yaxis.set_tick_params(labelsize=6)
        c.xaxis.set_tick_params(labelsize=8)
        c.grid(axis='x')

    def second_graph(self, b):
        b.set_title('Количество вакансий по годам', fontdict={'fontsize': 8})
        bar1 = b.bar(np.array(list(self.s2.keys())) - 0.4, self.s2.values(), width=0.4)
        bar2 = b.bar(np.array(list(self.s2.keys())), self.s4.values(), width=0.4)
        b.legend((bar1[0], bar2[0]), ('Количество вакансий', 'Количество вакансий\n' + self.vacancy.lower()),
                 prop={'size': 8})
        b.set_xticks(np.array(list(self.s2.keys())) - 0.2, list(self.s2.keys()), rotation=90)
        b.grid(axis='y')
        b.xaxis.set_tick_params(labelsize=8)
        b.yaxis.set_tick_params(labelsize=8)

    def first_graph(self, a):
        bar1 = a.bar(np.array(list(self.s1.keys())) - 0.4, self.s1.values(), width=0.4)
        bar2 = a.bar(np.array(list(self.s1.keys())), self.s3.values(), width=0.4)
        a.set_title('Уровень зарплат по годам', fontdict={'fontsize': 8})
        a.grid(axis='y')
        a.legend((bar1[0], bar2[0]), ('средняя з/п', 'з/п ' + self.vacancy.lower()), prop={'size': 8})
        a.set_xticks(np.array(list(self.s1.keys())) - 0.2, list(self.s1.keys()), rotation=90)
        a.xaxis.set_tick_params(labelsize=8)
        a.yaxis.set_tick_params(labelsize=8)

    def save(self, filename):
        self.book.save(filename=filename)

    def generate_pdf(self):
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        stats = []
        for year in self.s1.keys():
            stats.append([year, self.s1[year], self.s2[year], self.s3[year], self.s4[year]])

        for key in self.s6:
            self.s6[key] = round(self.s6[key] * 10 * 10, 2)

        pdf_template = template.render(
            {'name': self.vacancy, 'image_file': 'graph.png',
             'stats': stats, 'stats5': self.s5, 'stats6': self.s6})

        pdf_template = template.render(
            {'name': self.vacancy, 'path': '{0}/{1}'.format(pathlib.Path(__file__).parent.resolve(), 'graph.png'),
             'stats': stats, 'stats5': self.s5, 'stats6': self.s6})

        # config = pdfkit.configuration(wkhtmltopdf=r'/usr/bin/wkhtmltopdf')
        # pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})

        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})
        # pdfkit.from_string(pdf_template, 'report.pdf', options={"enable-local-file-access": ""})


InputConnect()
